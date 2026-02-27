# -*- coding: utf-8 -*-
"""数据文件读取模块 - 支持Excel和CSV"""

import os
import csv
from text_parser import extract_name_from_text, looks_like_person_name

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except Exception:
    load_workbook = None  # type: ignore
    HAS_OPENPYXL = False


def column_index_from_string(col):
    """兼容 openpyxl.utils.column_index_from_string（1-based）。"""
    s = str(col or "").strip().upper()
    if not s or not s.isalpha():
        raise ValueError(f"无效列标识: {col}")
    val = 0
    for ch in s:
        val = val * 26 + (ord(ch) - ord('A') + 1)
    return val


def get_column_letter(index):
    """兼容 openpyxl.utils.get_column_letter（1-based）。"""
    n = int(index or 0)
    if n <= 0:
        raise ValueError(f"无效列索引: {index}")
    out = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out.append(chr(ord('A') + rem))
    return "".join(reversed(out))


class DataReader:
    """数据文件读取器，支持Excel和CSV"""
    
    def __init__(self, filepath, name_col='E', intro_cols='F', url_col='G', start_row=2, source_col=None):
        """
        初始化数据读取器
        
        Args:
            filepath: 文件路径 (.xlsx, .xls, .csv)
            name_col: 姓名所在列 (如 'E' 或列索引1,2,3...)
            intro_cols: 简介所在列，支持多列 (如 'F' 或 'F,G,H' 或 '6,7,8')
            url_col: 图片链接所在列 (如 'G')
            start_row: 数据起始行 (默认2，跳过表头)
            source_col: 来源网页链接所在列 (可选)
        """
        self.filepath = filepath
        self.start_row = start_row
        self.data = []
        
        # 解析列设置
        self.name_col = self._parse_column(name_col)
        self.intro_cols = self._parse_columns(intro_cols)  # 支持多列
        self.url_col = self._parse_column(url_col)
        self.source_col = self._parse_column(source_col) if source_col else None
        
        # 检测文件类型
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ('.xlsx', '.xls'):
            self.file_type = 'excel'
        elif ext == '.csv':
            self.file_type = 'csv'
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
    
    def _parse_column(self, col):
        """解析列标识，支持字母(A,B,C)或数字(1,2,3)"""
        if isinstance(col, int):
            return col
        col = str(col).strip()
        if col.isdigit():
            return int(col)
        return column_index_from_string(col.upper())
    
    def _parse_columns(self, cols):
        """解析多列标识，支持逗号分隔"""
        if isinstance(cols, (list, tuple)):
            return [self._parse_column(c) for c in cols]
        
        cols_str = str(cols).strip()
        if ',' in cols_str:
            return [self._parse_column(c.strip()) for c in cols_str.split(',') if c.strip()]
        else:
            return [self._parse_column(cols_str)]
    
    def read(self):
        """读取数据"""
        if self.file_type == 'excel':
            return self._read_excel()
        else:
            return self._read_csv()
    
    def _read_excel(self):
        """读取Excel文件（含中文路径兼容处理）"""
        if not HAS_OPENPYXL or load_workbook is None:
            raise Exception("未安装 openpyxl，无法读取 Excel 文件。请安装 openpyxl 或改用 CSV。")

        self.data = []
        
        # 内部 Helper: 加载工作簿
        def load_wb(path):
            return load_workbook(path, read_only=True, data_only=True)

        wb = None
        temp_path = None
        
        try:
            file_ext = os.path.splitext(self.filepath)[1].lower()

            # 1. 尝试打开工作簿
            try:
                wb = load_wb(self.filepath)
            except Exception as e:
                if file_ext == '.xls':
                    raise Exception("不支持 .xls (旧版 Excel) 格式，请另存为 .xlsx 或导出为 .csv 后再导入。") from e
                # 失败回退：针对中文路径/编码问题，复制到临时文件尝试
                import tempfile
                import shutil
                try:
                    ext = os.path.splitext(self.filepath)[1]
                    fd, temp_path = tempfile.mkstemp(suffix=ext)
                    os.close(fd)
                    shutil.copy2(self.filepath, temp_path)
                    wb = load_wb(temp_path)
                except Exception:
                    # 如果临时文件也失败，则抛出原始异常
                    raise e
            
            # 2. 读取数据
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=self.start_row), start=self.start_row):
                name = self._get_cell_value(row, self.name_col)
                intro = self._get_intro_value(row, self.intro_cols)
                url = self._get_cell_value(row, self.url_col)
                source = self._get_cell_value(row, self.source_col) if self.source_col else None
                
                name_val = str(name).strip() if name is not None else ''
                url_val = str(url).strip() if url is not None else ''
                if not url_val:
                    continue

                # 兜底：抓取表格里的“题头/标题”常不可靠，优先从简介语义中抽取姓名
                intro_text = str(intro).strip() if intro else ''
                derived = extract_name_from_text(intro_text)
                if not name_val:
                    if not derived:
                        continue
                    name_val = derived
                elif derived and looks_like_person_name(derived):
                    # 如果原“姓名列”值不像人名，或不出现在简介开头，则以简介语义抽取为准
                    if (not looks_like_person_name(name_val)) or (intro_text and not intro_text.startswith(name_val)):
                        name_val = derived
                
                item = {
                    'row': row_num,
                    'name': name_val,
                    'intro': intro,
                    'url': url_val
                }
                if source:
                    item['source'] = str(source).strip()
                self.data.append(item)
                
        except Exception as e:
            print(f"读取Excel失败: {e}")
            raise e
            
        finally:
            if wb:
                try: wb.close()
                except: pass
            if temp_path and os.path.exists(temp_path):
                try: os.remove(temp_path)
                except: pass
        
        return self.data
    
    def _read_csv(self):
        """读取CSV文件"""
        self.data = []
        
        try:
            # 尝试检测编码
            encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030']
            content = None
            
            for encoding in encodings:
                try:
                    with open(self.filepath, 'r', encoding=encoding) as f:
                        content = f.read()
                    break
                except UnicodeDecodeError:
                    continue
            
            if content is None:
                raise Exception("无法识别文件编码")
            
            # 解析CSV
            lines = content.splitlines()
            reader = csv.reader(lines)
            
            for row_num, row in enumerate(reader, start=1):
                if row_num < self.start_row:
                    continue
                
                max_col = max(self.name_col, self.url_col, *self.intro_cols)
                if self.source_col:
                    max_col = max(max_col, self.source_col)
                if len(row) < max_col:
                    continue
                
                name = row[self.name_col - 1] if self.name_col <= len(row) else None
                intro = self._get_intro_from_list(row, self.intro_cols)
                url = row[self.url_col - 1] if self.url_col <= len(row) else None
                source = row[self.source_col - 1] if self.source_col and self.source_col <= len(row) else None
                
                name_val = str(name).strip() if name is not None else ''
                url_val = str(url).strip() if url is not None else ''
                if not url_val:
                    continue

                # 兜底：抓取表格里的“题头/标题”常不可靠，优先从简介语义中抽取姓名
                intro_text = str(intro).strip() if intro else ''
                derived = extract_name_from_text(intro_text)
                if not name_val:
                    if not derived:
                        continue
                    name_val = derived
                elif derived and looks_like_person_name(derived):
                    # 如果原“姓名列”值不像人名，或不出现在简介开头，则以简介语义抽取为准
                    if (not looks_like_person_name(name_val)) or (intro_text and not intro_text.startswith(name_val)):
                        name_val = derived
                
                item = {
                    'row': row_num,
                    'name': name_val,
                    'intro': intro,
                    'url': url_val
                }
                if source:
                    item['source'] = str(source).strip()
                self.data.append(item)
                
        except Exception as e:
            raise Exception(f"读取CSV文件失败: {str(e)}")
        
        return self.data
    
    def _get_cell_value(self, row, col_index):
        """获取指定列的单元格值"""
        try:
            cell = row[col_index - 1]
            return cell.value
        except IndexError:
            return None
    
    def _get_intro_value(self, row, col_indices):
        """获取多列简介并合并"""
        parts = []
        for col_idx in col_indices:
            val = self._get_cell_value(row, col_idx)
            if val:
                parts.append(str(val).strip())
        return ' | '.join(parts) if parts else ''
    
    def _get_intro_from_list(self, row, col_indices):
        """从列表中获取多列简介"""
        parts = []
        for col_idx in col_indices:
            if col_idx <= len(row):
                val = row[col_idx - 1]
                if val:
                    parts.append(str(val).strip())
        return ' | '.join(parts) if parts else ''
    
    def get_count(self):
        """返回数据条数"""
        return len(self.data)


def preview_raw_table(filepath, max_rows=100):
    """
    预览原始表格数据（用于用户选择列）
    
    Args:
        filepath: 文件路径
        max_rows: 最大预览行数
    
    Returns:
        dict: {
            'headers': ['A', 'B', 'C', ...],  # 列标识
            'data': [[row1], [row2], ...],    # 原始数据
            'total_rows': int,                 # 总行数
            'total_cols': int                  # 总列数
        }
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext in ('.xlsx', '.xls'):
        return _preview_excel(filepath, max_rows)
    elif ext == '.csv':
        return _preview_csv(filepath, max_rows)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _preview_excel(filepath, max_rows):
    """预览Excel文件"""
    if not HAS_OPENPYXL or load_workbook is None:
        raise Exception("未安装 openpyxl，无法预览 Excel 文件。请安装 openpyxl 或改用 CSV。")
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        data = []
        total_rows = 0
        max_col = 0
        
        for row_num, row in enumerate(ws.iter_rows(), start=1):
            total_rows += 1
            
            # 获取这一行的数据
            row_data = []
            for cell in row:
                val = cell.value
                row_data.append(str(val) if val is not None else '')
            
            # 更新最大列数
            if len(row_data) > max_col:
                max_col = len(row_data)
            
            if row_num <= max_rows:
                data.append(row_data)
        
        wb.close()
        
        # 生成列标识 (A, B, C, ...)
        headers = [get_column_letter(i) for i in range(1, max_col + 1)]
        
        # 确保每行都有相同列数
        for row in data:
            while len(row) < max_col:
                row.append('')
        
        return {
            'headers': headers,
            'data': data,
            'total_rows': total_rows,
            'total_cols': max_col
        }
        
    except Exception as e:
        if os.path.splitext(filepath)[1].lower() == '.xls':
            raise Exception("不支持 .xls (旧版 Excel) 格式，请另存为 .xlsx 或导出为 .csv 后再导入。") from e
        raise Exception(f"预览Excel文件失败: {str(e)}")


def _preview_csv(filepath, max_rows):
    """预览CSV文件"""
    try:
        # 尝试检测编码
        encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030']
        content = None
        
        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise Exception("无法识别文件编码")
        
        lines = content.splitlines()
        reader = csv.reader(lines)
        
        data = []
        max_col = 0
        total_rows = 0
        
        for row_num, row in enumerate(reader, start=1):
            total_rows += 1
            if len(row) > max_col:
                max_col = len(row)
            if row_num <= max_rows:
                data.append(row)
        
        # 生成列标识
        headers = [get_column_letter(i) for i in range(1, max_col + 1)]
        
        # 确保每行都有相同列数
        for row in data:
            while len(row) < max_col:
                row.append('')
        
        return {
            'headers': headers,
            'data': data,
            'total_rows': total_rows,
            'total_cols': max_col
        }
        
    except Exception as e:
        raise Exception(f"预览CSV文件失败: {str(e)}")


def read_data(filepath, name_col='E', intro_cols='F', url_col='G', start_row=2):
    """
    便捷函数：读取数据文件
    """
    reader = DataReader(filepath, name_col, intro_cols, url_col, start_row)
    return reader.read()


def auto_detect_columns(filepath):
    """
    自动检测数据文件中的列（姓名、说明、图片URL、来源URL）
    
    Args:
        filepath: 文件路径
    
    Returns:
        dict: {
            'name_col': str,       # 姓名列（如 'E'）
            'intro_cols': str,     # 说明列（如 'F' 或 'F,G'）
            'url_col': str,        # 图片URL列（如 'G'）
            'source_col': str,     # 来源URL列（如 'H'，可能为空）
            'start_row': int,      # 起始行（1=有表头时为2，无表头时为1）
            'confidence': str      # 检测置信度（high/medium/low）
        }
    """
    import re
    from text_parser import looks_like_person_name
    
    try:
        preview = preview_raw_table(filepath, max_rows=20)
    except Exception:
        return None
    
    if not preview or not preview['data']:
        return None
    
    headers = preview['headers']
    data = preview['data']
    total_cols = preview['total_cols']
    
    # 检测结果
    result = {
        'name_col': '',
        'intro_cols': '',
        'url_col': '',
        'source_col': '',
        'start_row': 2,
        'confidence': 'low'
    }
    
    # 关键词匹配（用于表头识别）
    # 说明：抓取表格里 “title/标题” 往往是网页题头，不一定是人名，权重应当很低
    name_keywords = ['姓名', '名字', '名称', 'name', '人名', 'person']
    weak_name_keywords = ['标题', 'title']
    intro_keywords = ['简介', '说明', '描述', '介绍', 'description', 'intro', 'desc', '信息', '详情', '备注']
    url_keywords = ['图片', '链接', 'url', 'image', 'img', '头像', '照片', 'photo', 'pic']
    source_keywords = ['来源', '网址', 'source', 'link', '原文', '原址', '详情链接', 'page']
    
    # URL模式
    url_pattern = re.compile(r'^https?://', re.IGNORECASE)
    image_url_pattern = re.compile(r'\.(jpg|jpeg|png|gif|webp|bmp)(\?|$)', re.IGNORECASE)
    
    # 分析每一列
    col_scores = []
    first_row = data[0] if data else []
    
    # 检测是否有表头（第一行是否像表头）
    has_header = False
    if first_row:
        # 如果第一行包含关键词，可能是表头
        first_row_text = ' '.join(str(c).lower() for c in first_row)
        all_keywords = name_keywords + intro_keywords + url_keywords + source_keywords
        if any(kw.lower() in first_row_text for kw in all_keywords):
            has_header = True
        # 如果第一行没有URL但后面行有URL，可能是表头
        elif data and len(data) > 1:
            first_has_url = any(url_pattern.match(str(c)) for c in first_row if c)
            second_has_url = any(url_pattern.match(str(c)) for c in data[1] if c)
            if not first_has_url and second_has_url:
                has_header = True
    
    result['start_row'] = 2 if has_header else 1
    data_rows = data[1:] if has_header else data
    
    for col_idx in range(total_cols):
        col_letter = headers[col_idx]
        header_val = str(first_row[col_idx]).lower() if col_idx < len(first_row) else ''
        
        score = {
            'idx': col_idx,
            'col': col_letter,
            'name_score': 0,
            'intro_score': 0,
            'url_score': 0,
            'source_score': 0
        }
        
        # 1. 表头关键词匹配（低权重；以内容语义为主）
        if has_header:
            for kw in name_keywords:
                if kw.lower() in header_val:
                    score['name_score'] += 3
            for kw in weak_name_keywords:
                if kw.lower() in header_val:
                    score['name_score'] += 1
            for kw in intro_keywords:
                if kw.lower() in header_val:
                    score['intro_score'] += 3
            for kw in url_keywords:
                if kw.lower() in header_val:
                    score['url_score'] += 3
            for kw in source_keywords:
                if kw.lower() in header_val:
                    score['source_score'] += 3
        
        # 2. 内容模式分析
        url_count = 0
        image_url_count = 0
        short_text_count = 0
        long_text_count = 0
        name_like_count = 0
        non_empty_count = 0
        
        for row in data_rows[:10]:  # 分析前10行数据
            if col_idx >= len(row):
                continue
            val = str(row[col_idx]).strip()
            if not val:
                continue
            non_empty_count += 1

            if looks_like_person_name(val):
                name_like_count += 1
            
            # 检查是否是URL
            if url_pattern.match(val):
                url_count += 1
                if image_url_pattern.search(val):
                    image_url_count += 1
            # 检查文本长度
            elif len(val) <= 20:
                short_text_count += 1
            else:
                long_text_count += 1
        
        if non_empty_count > 0:
            name_like_ratio = name_like_count / non_empty_count

            # 姓名列（语义识别）：人名比例越高分越高
            if name_like_ratio >= 0.6:
                score['name_score'] += 12
            elif name_like_ratio >= 0.35:
                score['name_score'] += 8
            elif name_like_ratio >= 0.2:
                score['name_score'] += 5

            # 图片URL列：高比例的图片URL
            if image_url_count / non_empty_count > 0.5:
                score['url_score'] += 8
            # 来源URL列：有URL但不是图片URL
            elif url_count / non_empty_count > 0.5 and image_url_count < url_count:
                score['source_score'] += 8
            # 姓名列：短文本为主（弱特征，避免被“题头/岗位”误导）
            if short_text_count / non_empty_count > 0.7 and url_count == 0:
                score['name_score'] += 3
            # 说明列：长文本为主
            if long_text_count / non_empty_count > 0.3 and url_count == 0:
                score['intro_score'] += 5
        
        col_scores.append(score)

    # 3. 语义关联：人名列常常出现在简介开头（“张三，男...”）
    # 找到最可能的简介列，然后提升“其值出现在简介开头”的列分数。
    best_intro = max(col_scores, key=lambda s: s.get('intro_score', 0), default=None)
    intro_idx = best_intro.get('idx') if best_intro and best_intro.get('intro_score', 0) > 0 else None
    if intro_idx is not None:
        for s in col_scores:
            if s.get('idx') == intro_idx:
                continue

            matches = 0
            total = 0
            for row in data_rows[:10]:
                if s['idx'] >= len(row) or intro_idx >= len(row):
                    continue

                name_val = str(row[s['idx']]).strip() if row[s['idx']] is not None else ''
                intro_val = str(row[intro_idx]).strip() if row[intro_idx] is not None else ''
                if not name_val or not intro_val:
                    continue

                if not looks_like_person_name(name_val):
                    continue

                total += 1
                # 强匹配：简介以人名起头
                if intro_val.startswith(name_val) or re.match(rf"^{re.escape(name_val)}[，,、\s]", intro_val):
                    matches += 1

            if total > 0:
                ratio = matches / total
                if ratio >= 0.6:
                    s['name_score'] += 8
                elif ratio >= 0.3:
                    s['name_score'] += 4
    
    # 根据分数选择最佳列
    def find_best_col(score_key, exclude=[]):
        best_col = ''
        best_score = 0
        for s in col_scores:
            if s['col'] not in exclude and s[score_key] > best_score:
                best_score = s[score_key]
                best_col = s['col']
        return best_col, best_score
    
    used_cols = []
    
    # 优先：图片URL列（最容易识别）
    url_col, url_score = find_best_col('url_score', used_cols)
    if url_col and url_score > 0:
        result['url_col'] = url_col
        used_cols.append(url_col)
    
    # 其次：来源URL列
    source_col, source_score = find_best_col('source_score', used_cols)
    if source_col and source_score > 0:
        result['source_col'] = source_col
        used_cols.append(source_col)
    
    # 姓名列
    name_col, name_score = find_best_col('name_score', used_cols)
    if name_col and name_score > 0:
        result['name_col'] = name_col
        used_cols.append(name_col)
    
    # 说明列（可能是多列）
    intro_cols = []
    for s in col_scores:
        if s['col'] not in used_cols and s['intro_score'] > 0:
            intro_cols.append(s['col'])
    if intro_cols:
        result['intro_cols'] = ','.join(intro_cols)
    
    # 计算置信度
    total_score = url_score + name_score
    if total_score >= 15:
        result['confidence'] = 'high'
    elif total_score >= 8:
        result['confidence'] = 'medium'
    else:
        result['confidence'] = 'low'
    
    return result
